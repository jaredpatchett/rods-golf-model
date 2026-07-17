#!/usr/bin/env python3
"""
Rod's Golf Model — course_fit_template.xlsx -> course_fit_table.json
=======================================================================
Right now course_fit_template.xlsx is pure reference: you fill it in, but
nothing reads it. This script closes that gap — it reads the "Courses" sheet
and writes course_fit_table.json, which rods_pipeline.py now loads
automatically (falling back to the hardcoded Birkdale/Bay Hill rows only if
this file doesn't exist).

Requires openpyxl (the one non-stdlib dependency in this project, and only
for this offline conversion step — the live pipeline itself stays stdlib-only):
    pip install openpyxl

Usage:
    python3 load_course_fit_xlsx.py course_fit_template.xlsx
    # -> writes course_fit_table.json next to it

Row 1 = column headers, row 2 = tier labels (skipped), row 3 = the example
row (skipped — it's Bay Hill, marked as an example, not real event data),
row 4+ = your actual courses.
"""

import sys
import json

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency. Run:  pip install openpyxl --break-system-packages")

EXAMPLE_COURSE_ID = "bay_hill"  # the template's built-in example row; always skipped


def load(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Courses"] if "Courses" in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    courses = []
    for row in rows[1:]:
        if row is None or all(v is None for v in row):
            continue
        record = dict(zip(headers, row))
        cid = record.get("course_id")
        if not cid or cid == EXAMPLE_COURSE_ID or str(cid).lower().startswith("example"):
            continue
        # tier-label row ("TIER 1 pull", "id/derived", ...) has no numeric course_id row shape;
        # cheapest real check is whether course_name is a plausible string, not a label
        if record.get("course_name") in (None, "TIER 1 pull", "id/derived", "TIER 2 infer", "TIER 3 manual"):
            continue
        courses.append(record)
    return courses


def to_F_row(course):
    """
    Map a filled course_fit_template.xlsx row -> the page's F-row shape:
    [course, renewals, scoreAdj, scoreSD, blowup, wOTT,wAPP,wARG,wPUT, distPrem, recovPen, cutParar, grass]
    Falls back to None for any field left blank; the HTML page renders those
    cells as best it can (heat scaling skips nulls already in the existing JS).
    """
    return [
        course.get("course_name"),
        course.get("n_renewals"),
        course.get("scoring_avg_adj"),
        course.get("scoring_sd_adj"),
        course.get("skew_blowup_rate"),
        course.get("w_ott"),
        course.get("w_app"),
        course.get("w_arg"),
        course.get("w_put"),
        course.get("dist_premium"),
        course.get("recovery_penalty"),
        course.get("cutline_vs_par"),
        course.get("green_grass"),
    ]


def main():
    if len(sys.argv) < 2:
        sys.exit("Usage: python3 load_course_fit_xlsx.py course_fit_template.xlsx")
    path = sys.argv[1]
    courses = load(path)
    if not courses:
        print("No filled-in course rows found (only the example row was present). "
              "Nothing written — rods_pipeline.py will keep using its hardcoded fallback.")
        return
    full_records = courses                       # keep every column, for course_fit_engine.py use
    f_rows = [to_F_row(c) for c in courses]       # page-ready rows, for rods_data.json's F array
    out = {"courses": full_records, "F": f_rows}
    out_path = "course_fit_table.json"
    with open(out_path, "w") as f:
        json.dump(out, f, indent=1)
    print(f"Wrote {out_path}: {len(courses)} course(s).")


if __name__ == "__main__":
    main()
